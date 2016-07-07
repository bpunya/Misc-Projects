import os
import math
import string
import openpyxl
import time
import tkinter.filedialog

################################################################################
class Docket():
    # This class is the structure of the docketrequirements object.
    # Docket.name is a string containing the name of the docket.
    # Docket.requirements is a list of all cell codes that the docket includes

    def __init__(self, name):
        self.name = name
        self.requirements = []


################################################################################
#   programStart() is the only function directly called by the GUI             #
################################################################################
def programStart(self): #self
    def msg(input):
        # Make the message change function easier to access.
        self.parent.parent.workingmessage.set(input)
    def frame(input):
        # Make the changeFrame function easier to access.
        self.parent.parent.changeFrame(input)

    debug = False       # Set debug to true to ignore exception checks (only use with preset files)
    logdebug = False    # Set logdebug to true to log docket and cellcount objects to console.

    msg("Please wait, the application is running.")
    frame("MSG")

    template = ""
    template = self.parent.templateinput.selectedfilename.get()
    inputslist = []
    inputslist = list(self.parent.inputlist)

    # Doubles of the variables are listed above for easy debugging.
    # Comment out the second instance of the variable to use the preset filelist for debugging.

    docketrequirements = []     # docketrequirements is a list of Docket objects (see above) that contain a list of cell codes in each active docket.
    cellcounts = []             # cellcounts is a list of cellcount dictionaries that contain the quantities of cellcodes mailed in each cycle
    filelist = []               # filelist is a collection of all files input into the GUI. The first item should always be the template.

    # Check to see if a template file was included. If not, send the user an error message and stop the program.
    # All other error catching functions include the same two msg("ERROR!"), frame("BACK") lines.

    if len(template) < 1 and debug == False:
        msg("No template file was selected.")
        frame("BACK")
        return
    elif debug == False:
        filelist.append(template)

    # Add all inputs given that are real to the filelist list.
    # If you want to do more validation for inputs, such as ensuring that users have access to the files, this is where it would happen.

    for rawinput in inputslist:
        try:
            if os.path.isfile(rawinput) and debug == False:
                item = rawinput
                filelist.append(item)
        except:
            msg("An error has occured when reading the selected files")
            frame("BACK")
            return

    # If debugging is active, use a preset list of files (so you don't have to run the program each time)
    # The first file should always be the template file. The rest should all be valid files. Make sure to properly escape the backslashes.

    if debug:
        filelist = []
    if len(filelist) < 2:
        msg("There were no files input.")
        frame("BACK")
        return

    # The first major parsing function. Parses only the template file for docket information and returns it into the docketrequirements object.
    # The programParseTemplate function returns a list of Docket objects. On error, stop the program and return the user.

    try:
        docketrequirements = programParseTemplate(filelist[0])
    except:
        msg("An error has occured when parsing the template file.")
        frame("BACK")
        return
    if docketrequirements == False:
        msg("You do not have permission to read " + filelist[0])
        frame("BACK")
        return

    # In case the files processed are very large or numerous, this is a message to list the dockets being processed by the program.
    # This will return all dockets listed (properly) in the docket template. If this screen shows incorrect dockets, then there is a problem with the template file.

    update = ""
    for item in docketrequirements:
        update = update + item.name + ", "
    msg('''Please wait, the application is running.
The following Dockets are being processed (nothing saved yet):

'''+ update)

    # programParseLogs is run one at a time for each log file that is entered. This is so that problem files can be correctly error reported and so count information can be separated.
    # The programParseLogs function returns a single dictionary that contains only one cycle's worth of data. The format is as follows:
    #    cellcount = {
    #                   'name' = LOGNAME
    #                   'cellcode1' = QUANTITY
    #                   'cellcode2' = QUANTITY
    #                 }

    for item in filelist[1:]:
        currentfile = str(item)
        try:
            singlelog = programParseLogs(item)
            cellcounts.append(dict(singlelog))
        except:
            msg('''An error has occured when parsing a file. File location:

'''+ currentfile)
            frame("BACK")
            return

        if singlelog == False:
            msg("You do not have permission to read " + currentfile)
            frame("BACK")
            return

    # Logging should be done here. This is where data collection ends and file creation begins.
    # Add any other logging functions below if you want to check the ability for the program to properly parse template files.

    if logdebug == True:
        programDebug(docketrequirements, cellcounts)

    # programCreateExcelFile handles all spreadsheet creation. It takes all the docket information and combines them in a
    # single workbook. It CANNOT update existing sheets and so the information inside should be moved to an archivable file

    try:
        docketfile = programCreateExcelFile(docketrequirements, cellcounts)
    except:
        msg("An error has occured when creating the excel files.")
        frame("BACK")
        return

    # Saving is the last step. Due to the difficulty of handling and saving multiple files, the spreadsheets are held in a single workbook file.
    # The docket information is held in separate worksheets in the workbook.

    msg("Please save your Excel Workbook. The default name is \"Generated on YYYY-MM-DD\"")
    try:
        programSaveExcelFile(docketfile)
    except:
        msg("An error has occured while the file was being saved.")
        frame("BACK")
        return

    msg("Program Complete. Press here to return to the main screen.")
    frame("BACK")

    ##################
    # END OF PROGRAM #
    ##################

################################################################################
#   safeOpen()                                                                 #
#                                                                              #
#   This is a safe open() function that checks for user authentication         #
################################################################################
def safeOpen(filelocation, opentype):
    if os.access(filelocation, os.R_OK):
        return open(filelocation, opentype)
    else:
        return False


################################################################################
#   programDebug()                                                             #
#                                                                              #
#   This is a simple debug function that prints all objects in use to the      #
#   Python Console.                                                            #
################################################################################
def programDebug(docketrequirements, cellcounts):
    print("CELL CODES IN EACH DOCKET:")
    for item in docketrequirements:
        print("\n" + item.name)
        for value in item.requirements:
            print(value)
    print("\n CELL COUNT DATA:")
    for item in cellcounts:
        print("\n" + str(item))


################################################################################
#   programSaveExcelFile()                                                     #
#                                                                              #
#   This generates file details and handles saving.                            #
################################################################################
def programSaveExcelFile(docket): #requires docket

    defaultname = "Generated " + time.strftime("%Y-%m-%d") + " at " + time.strftime("%I%p") + ".xlsx"
    excel = [('Microsoft Excel 2007-2013 XML', '.xlsx')]
    docket.save(tkinter.filedialog.asksaveasfilename(filetypes=excel, initialfile=defaultname))


################################################################################
#   programParseTemplate()                                                     #
#                                                                              #
#   This function parses the template file given for docket information. The   #
#   template file must have a blank line after the end of a docket block or    #
#   it WILL NOT find the last block.                                           #
################################################################################
def programParseTemplate(filelocation):

    rawfile = safeOpen(filelocation, 'r')
    ## If no access, cancel function ##
    if rawfile == False:
        return False
    docketrequirements = []
    docketcodelist = []
    filldocket = False

    for i,rawline in enumerate(rawfile):
        line = rawline.replace('\n','')

        if filldocket and bool(line != ""):
            docketcodelist.append(line)
        elif filldocket and bool(line == ""):
            filldocket = not filldocket
            docket = docketcodelist[0]
            docket.requirements = docketcodelist[1:]
            docketrequirements.append(docket)
            docketcodelist = []
        elif line.split(" ")[0].upper() == "DOCKET":
            filldocket = not filldocket
            newdocket = Docket(line)
            docketcodelist.append(newdocket)

    return docketrequirements


################################################################################
#   programParseLogs()                                                         #
#                                                                              #
#   This function returns a list of dictionaries containing cell code counts.  #
#   The way it collects data is too rigid. Regex is suggested.                 #
################################################################################
def programParseLogs(filelocation):
    filename = os.path.basename(filelocation)
    log = {}
    # MUST CHANGE WHEN DIFFERENT LOG FILES ARE INPUT #
    log['name'] = filename[-18:-4]

    # Needs a way to verify if this is a log file #
    rawfile = safeOpen(filelocation, 'r')
    # If access to the file is not allowed, cancel function #
    if rawfile == False:
        return False

    # For loop parses the text file line by line #
    for i, rawline in enumerate(rawfile):
        line = rawline.replace('\n','')

        # MUST CHANGE WHEN NEW FILES ARE ADDED #
        if str(line[16:19]).upper() == "QUA":
            cellcode = line[11:15]
            rawquantity = line[21:]
            quantity = rawquantity.lstrip("0")
            if cellcode in log:
                log[cellcode] = log[cellcode] + quantity
            else:
                log[cellcode] = quantity

    return sorted(log.items())


################################################################################
#   cell()                                                                     #
#                                                                              #
#   This function returns a cell position as a string. The default cell        #
#   returned with an input of (0,0) is "A4".                                   #
################################################################################
def cell(column, row):
    if column <= 25:
        columnvalue = string.ascii_uppercase[column]
    else:
        firstcharacter = int(math.floor(column / 26))-1
        secondcharacter = int(column % 26)
        columnvalue = string.ascii_uppercase[firstcharacter]+string.ascii_uppercase[secondcharacter]

    rowvalue = str(row + 4)
    return str(columnvalue + rowvalue)


################################################################################
#   programCreateExcelFile()                                                   #
#                                                                              #
#   This function returns a single Workbook object that has a sheet for every  #
#   dockets listed in the docket template file. A single Workbook is returned. #
################################################################################
def programCreateExcelFile(docketrequirements, cellcounts): #docketrequirements, cellcounts

    workbook = openpyxl.Workbook()
    infosheet = workbook.active

    # Set up workbook defaults. This is done every time.
    infosheet.title = "Read Me"
    infosheet['B2'] = "This workbook was automatically generated."
    infosheet['B3'] = "It is advised that you copy and paste from this workbook into the proper files, instead of using this as a base."
    infosheet['B4'] = "The program that was used to create this file cannot update existing files."

    ### Create a worksheet to host all unknown values. It will be used later ###
    unknowns = workbook.create_sheet()
    unknowns.title = "Extras"
    unknowns['A1'] = "Unknown Cell Codes"
    unknowns['A3'] = "Cell Code"
    unknowns['B3'] = "Quantity"
    unknowns['C3'] = "From Cycle"
    unknownsheetrow = 0

    # Create a worksheet for every docket listed in the docket requirements template.
    for docket in docketrequirements:
        worksheet = workbook.create_sheet()
        worksheet.title = docket.name
        worksheet['A1'] = docket.name
        worksheet['A3'] = "Cell Code"

        # Set maximum width and height for styling purposes (totals).
        maxwidth = len(cellcounts)+1
        maxheight = len(docket.requirements)+1

        ### This function builds worksheets row by row.             ###
        for i, cellcode in enumerate(docket.requirements):
            ### It selects the first cell code listed in the docket ###
            ### requirements file and Lists it in the A column...   ###
            worksheet[cell(0,i)] = cellcode

            ### And then iterates through every cycle file...       ###
            for v, cyclefile in enumerate(cellcounts):
                ### Iterate v for stylistic purposes                ###
                v += 1
                ### Place the name of the current cycle on row 3    ###
                worksheet[cell(v,-1)] = cyclefile['name']

                ### If the cellcode exists in the selected cycle file, ###
                ### add the quantity into the proper cell.             ###
                if cellcode in cyclefile:
                    worksheet[cell(v,i)] = int(cyclefile[cellcode])
                    ### To separate the known from unknown quantities  ###
                    cyclefile[cellcode] = "PARSED"
                else:
                    worksheet[cell(v,i)] = 0

            ### While the function is still iterating left to right, add a totals column to the right of the last object. ###
            worksheet[cell(maxwidth,i)] = "=SUM("+cell(1,i)+":"+cell(maxwidth-1,i)+")"

        ### This adds the bottom totals row ###
        for x in range(0,maxwidth+1):
            worksheet[cell(x,maxheight)] = "=SUM("+cell(x,0)+":"+cell(x,i)+")"

        ### Descriptor cells are added here, after you know the totals of everything. ###
        worksheet[cell(0,maxheight)] = "Cycle Total"
        worksheet[cell(maxwidth,-1)] = "Total Items Mailed"

    ### Parse the cycle files again to find all that were not captured by the first sweep.  ###
    for cyclefile in cellcounts:
        for i, cellcode in enumerate(cyclefile):
            if not cyclefile[cellcode] == "PARSED" and not cellcode == 'name':
                unknowns[cell(0,unknownsheetrow)] = cellcode
                unknowns[cell(1,unknownsheetrow)] = int(cyclefile[cellcode])
                unknowns[cell(2,unknownsheetrow)] = cyclefile['name']
                unknownsheetrow += 1


    return workbook


################################################################################
